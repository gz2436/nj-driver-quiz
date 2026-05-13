#!/usr/bin/env python3
"""Export data/questions.json -> data/questions_for_cleaning.xlsx.

One row per question. Bilingual text is auto-split into _en / _zh columns
at the first CJK transition (best-effort, manual cleanup expected).

Workflow:
  1. python3 tools/export_for_cleaning.py    # produces xlsx
  2. User cleans columns in Excel/Numbers
  3. python3 tools/import_cleaned.py         # writes back into questions.json
     (importer to be written after user finishes cleaning)
"""
import json
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path('/Users/gavincheung/NYU/Driver')
DATA = ROOT / 'data'

CJK_RE = re.compile(r'[一-鿿]')


def split_bilingual(text):
    """Split 'English text 中文' into ('English text', '中文').

    If there's no CJK at all, returns (text, ''). If text is pure CJK,
    returns ('', text). Splits at the FIRST CJK character.
    """
    if not text:
        return '', ''
    m = CJK_RE.search(text)
    if not m:
        return text.strip(), ''
    en = text[:m.start()].strip()
    zh = text[m.start():].strip()
    return en, zh


def main():
    qs = json.loads((DATA / 'questions.json').read_text())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'questions'

    headers = [
        'id', 'type', 'source_doc', 'source_para_idx',
        'stem_en', 'stem_zh',
        'A_en', 'A_zh',
        'B_en', 'B_zh',
        'C_en', 'C_zh',
        'D_en', 'D_zh',
        'answer',
        'answer_text_en', 'answer_text_zh',
        'stem_img',
        'topics',
        'is_common_mistake',
        'is_duplicate_of',
        'needs_review',
        'verified',
        'explanation_key',
        'notes',  # blank — for user to add manual notes during cleaning
    ]
    ws.append(headers)

    # Header styling
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='374151')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    review_fill = PatternFill('solid', fgColor='FEF3C7')   # amber for needs_review
    dup_fill = PatternFill('solid', fgColor='F3F4F6')      # gray for duplicates
    qa_fill = PatternFill('solid', fgColor='DBEAFE')        # blue for QA

    for q in qs:
        opts = q.get('options', []) or []
        opts = opts + [''] * (4 - len(opts))
        a_en, a_zh = split_bilingual(opts[0])
        b_en, b_zh = split_bilingual(opts[1])
        c_en, c_zh = split_bilingual(opts[2])
        d_en, d_zh = split_bilingual(opts[3])
        s_en, s_zh = split_bilingual(q.get('stem', ''))
        at_en, at_zh = split_bilingual(q.get('answer_text') or '')

        row = [
            q['id'],
            q.get('type', 'mc'),
            q.get('source_doc', ''),
            q.get('source_para_idx', ''),
            s_en, s_zh,
            a_en, a_zh,
            b_en, b_zh,
            c_en, c_zh,
            d_en, d_zh,
            q.get('answer') or '',
            at_en, at_zh,
            q.get('stem_img') or '',
            ','.join(q.get('topics') or []),
            'Y' if q.get('is_common_mistake') else '',
            q.get('is_duplicate_of') or '',
            q.get('needs_review') or '',
            'Y' if q.get('verified') else '',
            q.get('explanation_key') or '',
            '',  # notes
        ]
        ws.append(row)

        row_idx = ws.max_row
        # Tint rows by status
        if q.get('needs_review'):
            for cell in ws[row_idx]:
                cell.fill = review_fill
        elif q.get('is_duplicate_of'):
            for cell in ws[row_idx]:
                cell.fill = dup_fill
        elif q.get('type', '').startswith('qa'):
            for cell in ws[row_idx]:
                cell.fill = qa_fill

    # Column widths
    widths = {
        'id': 6, 'type': 10, 'source_doc': 11, 'source_para_idx': 8,
        'stem_en': 50, 'stem_zh': 35,
        'A_en': 30, 'A_zh': 20, 'B_en': 30, 'B_zh': 20,
        'C_en': 30, 'C_zh': 20, 'D_en': 30, 'D_zh': 20,
        'answer': 7,
        'answer_text_en': 25, 'answer_text_zh': 25,
        'stem_img': 22,
        'topics': 18,
        'is_common_mistake': 6, 'is_duplicate_of': 8,
        'needs_review': 14, 'verified': 6,
        'explanation_key': 18, 'notes': 25,
    }
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 15)

    # Wrap text on long columns; freeze top row + id column
    wrap_cols = {'stem_en', 'stem_zh', 'A_en', 'A_zh', 'B_en', 'B_zh',
                 'C_en', 'C_zh', 'D_en', 'D_zh', 'answer_text_en', 'answer_text_zh', 'notes'}
    wrap_idx = {i + 1 for i, h in enumerate(headers) if h in wrap_cols}
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column in wrap_idx:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            else:
                cell.alignment = Alignment(vertical='top')

    ws.freeze_panes = 'B2'
    ws.auto_filter.ref = ws.dimensions

    # ---- Legend sheet ----
    ws2 = wb.create_sheet('legend')
    legend_rows = [
        ['Color', '含义 Meaning'],
        ['amber 琥珀色', 'needs_review — 待手工核对：可能少选项 / 答案缺失 / 格式异常'],
        ['blue 蓝色', 'Q&A 直答题：无 ABCD 选项，答案在 answer_text_zh 列'],
        ['gray 灰色', '重复题：is_duplicate_of 指向原题 id（保留供参考，可在 notes 里写"删"）'],
        ['', ''],
        ['Column', '说明'],
        ['stem_en / stem_zh', '题干英文 / 中文。自动按第一个汉字切。若切错，手工挪。'],
        ['A_en / A_zh ... D_en / D_zh', '4 个选项的英中分列。空格代表源里就缺。'],
        ['answer', '正确答案字母 A/B/C/D（mc 题用）'],
        ['answer_text_en / _zh', 'Q&A 题的直答内容'],
        ['needs_review', '导出时已知的格式问题；改好后可清空此列'],
        ['notes', '空白栏，你可以写"删"、"合并到 N"、"内容错"等指令，导入时我会按 notes 执行'],
        ['', ''],
        ['Workflow', ''],
        ['1. 用 Excel / Numbers / Google Sheets 打开'],
        ['2. 按 needs_review 列筛选先处理琥珀色'],
        ['3. 检查 stem_en/zh 切分对不对（自动按首个汉字切，少数会错位）'],
        ['4. 选项里 A_en/A_zh 内容对应是否正确'],
        ['5. 重复题（灰）想删就在 notes 写"删"'],
        ['6. 保存为 xlsx 给我，我写 importer 写回 questions.json'],
    ]
    for r in legend_rows:
        ws2.append(r)
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 80
    ws2['A1'].font = Font(bold=True)
    ws2['B1'].font = Font(bold=True)
    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    out = DATA / 'questions_for_cleaning.xlsx'
    wb.save(out)
    print(f'Wrote {out}')
    print(f'  Rows: {len(qs)}')
    print(f'  needs_review (amber): {sum(1 for q in qs if q.get("needs_review"))}')
    print(f'  Q&A (blue): {sum(1 for q in qs if q.get("type","").startswith("qa"))}')
    print(f'  Duplicates (gray): {sum(1 for q in qs if q.get("is_duplicate_of"))}')


if __name__ == '__main__':
    main()
