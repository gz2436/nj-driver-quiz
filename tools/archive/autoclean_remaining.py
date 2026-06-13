#!/usr/bin/env python3
"""Apply auto-fixes to rows 225+ in data/questions_for_cleaning.xlsx.

User has manually cleaned Q1-Q224 (rows 2-225). Leave those untouched.
For Q225+, apply two confidence-graded fixes:

  1. UNIT-NUMBER REPATRIATION (high confidence, auto-applied):
     If `X_en` ends with a bare number and `X_zh` starts with a CJK unit word
     (英尺/盎司/杯/倍/个月/天/年/英里/磅/美元/分钟/小时/$/...), move the number
     from the tail of en to the head of zh.

  2. CROSS-LINE OPTION LEAK (medium confidence, flagged only):
     Heuristics to detect when an option holds content that belongs to a
     different option or the next question. Mark needs_review = 'option_leak'.

Plus: detect "statement" style cards (option-less knowledge cards) and mark
type accordingly.

Output: data/questions_for_cleaning_v2.xlsx (preserves user's Q1-Q224 work).
"""
import re
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path('/Users/gavincheung/NYU/Driver')
DATA = ROOT / 'data'
SRC = DATA / 'questions_for_cleaning.xlsx'
OUT = DATA / 'questions_for_cleaning_v2.xlsx'

USER_DONE_THROUGH_ROW = 225  # rows 2..225 == Q1..Q224 (user-cleaned)

# Chinese unit words that often follow a bare number in source bilingual format
CJK_UNIT_HEAD = re.compile(
    r'^('
    r'英尺|英寸|英里|米|厘米|公里|公分|'           # length
    r'盎司|磅|公斤|克|加仑|公升|升|'                # weight/volume
    r'秒|分钟|小时|天|日|月|年|个月|周|星期|'        # time
    r'倍|次|次数|百分点|分|'                        # multiplier
    r'美元|元|分|'                                  # currency
    r'杯|瓶|罐|片|颗|粒|'                           # serving
    r'人|位|名|个|辆|条|项|'                       # countable noun
    r'度|摄氏度|华氏度|'
    r'mph|MPH|英里/小时|公里/小时|km/h'
    r')'
)

# Bare number pattern at end of en field
EN_TRAILING_NUMBER = re.compile(r'\s+(\d+(?:\.\d+)?)\s*$')

# Trailing money like "...text $500" or "...text $1,000.50"
EN_TRAILING_MONEY = re.compile(r'\s+(\$[\d,]+(?:\.\d+)?)\s*$')


def repatriate_unit_number(en_text, zh_text):
    """If en has trailing duplicated number/money that belongs in zh, move it.
    Returns (new_en, new_zh, did_change)."""
    if not en_text or not zh_text:
        return en_text, zh_text, False

    # Pattern 1: bare number at end of en + CJK unit at start of zh
    m = EN_TRAILING_NUMBER.search(en_text)
    if m and CJK_UNIT_HEAD.match(zh_text):
        num = m.group(1)
        new_en = en_text[:m.start()].rstrip(' \t')
        new_zh = num + zh_text
        return new_en, new_zh, True

    # Pattern 2: trailing $X duplicated (appears earlier in en too)
    m_money = EN_TRAILING_MONEY.search(en_text)
    if m_money:
        money_tok = m_money.group(1)            # e.g. "$500"
        normalized_num = money_tok.replace(',', '').lstrip('$')  # "500"
        # Check if the same amount (with/without comma) already appears earlier in en
        body = en_text[:m_money.start()]
        body_normalized = body.replace(',', '')
        if normalized_num in body_normalized or '$' + normalized_num in body_normalized:
            # Duplication: strip trailing token, prepend to zh (with space)
            new_en = body.rstrip(' \t')
            sep = '' if zh_text.startswith(('，', ',', '。')) else ' '
            new_zh = money_tok + sep + zh_text
            return new_en, new_zh, True

    return en_text, zh_text, False


# Statement-type heuristic markers found in user's manual cleanup
STATEMENT_MARKERS_EN = [
    'This is a',
    'This sign means',
    'This indicates',
]
STATEMENT_NO_OPTIONS_HINT = re.compile(r'(没有选项|无选项|陈述|flashcard|statement)', re.IGNORECASE)


def detect_statement(row_dict):
    """Detect statement-type cards (no real options, just a knowledge fact).

    Strict: must have ALL 4 options empty AND stem looks declarative
    ("This is a ___ sign" pattern). Avoids false positives on
    'Which statement is true?' MC questions.
    """
    a_en = (row_dict.get('A_en') or '').strip()
    if a_en == '（此题无选项）':
        return True

    opts_filled = sum(1 for L in 'ABCD'
                      if (row_dict.get(f'{L}_en') or '').strip()
                      or (row_dict.get(f'{L}_zh') or '').strip())
    if opts_filled > 0:
        return False

    stem_en = (row_dict.get('stem_en') or '').strip()
    stem_zh = (row_dict.get('stem_zh') or '').strip()

    if re.match(r'^This is an? .* (?:sign|road sign)\.?', stem_en, re.IGNORECASE):
        return True
    if re.search(r'路标|是一个.*标志', stem_zh):
        return True
    return False


def detect_option_leak(row_dict):
    """Detect option content that likely belongs elsewhere. Heuristic."""
    leaks = []
    # Patterns that look like the start of a NEW question (next-stem fragment)
    next_stem_signals = re.compile(
        r'(?:'
        r'A conviction of |Conviction for |The maximum |The minimum |'
        r'A driver who |A motorist |If you (?:are|have|will|may) |'
        r'When (?:driving|you are|you have|approaching) |'
        r'What (?:does|is) (?:this|the|a) |Which of the |'
        r'In New Jersey,? you must |On a (?:two|three|four)-lane '
        r')'
    )
    for L in 'ABCD':
        zh = (row_dict.get(f'{L}_zh') or '').strip()
        en = (row_dict.get(f'{L}_en') or '').strip()
        # 1. ZH option has a long English run after the Chinese (likely next-Q leak)
        if zh:
            # Find longest contiguous English run anywhere in zh
            eng_runs = re.findall(r'[A-Za-z][A-Za-z\s,\.\-\']{20,}', zh)
            if eng_runs:
                leaks.append(f'{L}_zh has long English run ({len(eng_runs[0])}+ chars)')
        # 2. EN option contains a next-stem signal phrase
        if en and next_stem_signals.search(en):
            leaks.append(f'{L}_en has next-question phrase')
        # 3. Option contains an inline letter-marker (e.g., "B. " mid-text)
        if en and re.search(r'\b[A-D]\.\s', en):
            leaks.append(f'{L}_en has inline-letter-marker')
    return leaks


def main():
    wb = openpyxl.load_workbook(SRC)
    ws = wb['questions']
    headers = [c.value for c in ws[1]]
    H = {h: i for i, h in enumerate(headers)}

    # Detect color marks the user added (red on id) to preserve them in v2
    red_marked = []
    for row_idx in range(2, USER_DONE_THROUGH_ROW + 1):
        id_cell = ws.cell(row=row_idx, column=H['id'] + 1)
        if id_cell.fill and id_cell.fill.fgColor:
            rgb = id_cell.fill.fgColor.rgb
            if rgb and isinstance(rgb, str):
                rgb = rgb.upper()
                if len(rgb) == 8:
                    r = int(rgb[2:4], 16); g = int(rgb[4:6], 16); b = int(rgb[6:8], 16)
                elif len(rgb) == 6:
                    r = int(rgb[0:2], 16); g = int(rgb[2:4], 16); b = int(rgb[4:6], 16)
                else:
                    continue
                if r > 200 and g < 120 and b < 120:
                    red_marked.append(row_idx)

    fixes_applied = []
    leaks_flagged = []
    statements_marked = []

    review_fill = PatternFill('solid', fgColor='FEF3C7')          # amber
    auto_fixed_fill = PatternFill('solid', fgColor='D1FAE5')      # green tint for our auto-fix
    statement_fill = PatternFill('solid', fgColor='E9D5FF')        # purple tint for statement

    # Iterate only over the AUTO-FIX region: rows USER_DONE_THROUGH_ROW+1 .. end
    for row_idx in range(USER_DONE_THROUGH_ROW + 1, ws.max_row + 1):
        row = ws[row_idx]
        qid = row[H['id']].value
        if not qid:
            continue

        row_dict = {h: (row[H[h]].value or '') for h in headers}

        # ----- Auto-fix 1: unit-number repatriation -----
        any_unit_fix = False
        for field_pair in [('stem_en', 'stem_zh'),
                           ('A_en', 'A_zh'),
                           ('B_en', 'B_zh'),
                           ('C_en', 'C_zh'),
                           ('D_en', 'D_zh'),
                           ('answer_text_en', 'answer_text_zh')]:
            en_field, zh_field = field_pair
            en_val = str(row_dict.get(en_field) or '')
            zh_val = str(row_dict.get(zh_field) or '')
            new_en, new_zh, changed = repatriate_unit_number(en_val, zh_val)
            if changed:
                row[H[en_field]].value = new_en
                row[H[zh_field]].value = new_zh
                row_dict[en_field] = new_en
                row_dict[zh_field] = new_zh
                any_unit_fix = True

        if any_unit_fix:
            fixes_applied.append(qid)

        # ----- Detect statement-type cards -----
        if detect_statement(row_dict):
            row[H['type']].value = 'statement'
            statements_marked.append(qid)
            # Color the whole row purple
            for cell in row:
                cell.fill = statement_fill
            continue  # don't double-color with leak flag

        # ----- Detect option-leak (flag only, don't change content) -----
        leaks = detect_option_leak(row_dict)
        if leaks:
            existing_nr = (row[H['needs_review']].value or '').strip()
            if 'option_leak' not in existing_nr:
                new_nr = (existing_nr + ';' if existing_nr else '') + 'option_leak'
                row[H['needs_review']].value = new_nr
            existing_notes = (row[H['notes']].value or '')
            hint = '; '.join(leaks)
            if hint not in existing_notes:
                row[H['notes']].value = (existing_notes + ' | ' if existing_notes else '') + f'auto-detected: {hint}'
            leaks_flagged.append((qid, leaks))
            for cell in row:
                cell.fill = review_fill
            continue

        # If row got an auto-fix (unit), tint green so user can sanity-check
        if any_unit_fix:
            # Only tint if no other amber/blue/purple status active
            existing_fill = row[H['stem_en']].fill.fgColor.rgb if row[H['stem_en']].fill and row[H['stem_en']].fill.fgColor else None
            if not existing_fill or existing_fill == '00000000':
                for cell in row:
                    cell.fill = auto_fixed_fill

    # Restore red fills on Q1-Q224 user-marked IDs
    red_fill = PatternFill('solid', fgColor='FCA5A5')
    for row_idx in red_marked:
        ws.cell(row=row_idx, column=H['id'] + 1).fill = red_fill

    # Add a small frozen note in cell that's not a row:
    # (skip — keep the file clean)

    wb.save(OUT)
    print(f'Wrote {OUT}')
    print(f'  User-preserved rows: 2..{USER_DONE_THROUGH_ROW} (Q1..Q{USER_DONE_THROUGH_ROW-1})')
    print(f'  Auto-scan range:    rows {USER_DONE_THROUGH_ROW+1}..{ws.max_row} (Q{USER_DONE_THROUGH_ROW}..Q{ws.max_row-1})')
    print(f'  Unit-number fixes:  {len(fixes_applied)} questions')
    print(f'  Option-leak flags:  {len(leaks_flagged)} questions')
    print(f'  Statement marks:    {len(statements_marked)} questions')
    print()
    print(f'  Red-marked IDs preserved: {len(red_marked)}')
    if fixes_applied[:10]:
        print(f'  Sample unit-fix IDs:    {fixes_applied[:10]}')
    if leaks_flagged[:5]:
        print(f'  Sample option-leak IDs: {[x[0] for x in leaks_flagged[:5]]}')
    if statements_marked:
        print(f'  Statement IDs:          {statements_marked}')


if __name__ == '__main__':
    main()
