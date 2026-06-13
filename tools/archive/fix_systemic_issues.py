#!/usr/bin/env python3
"""Fix systemic issues in data/questions_for_cleaning_v2.xlsx for Q225+.

Three patterns auto-fixed:

  1. STEM_AS_OPTION_A: A_zh holds Chinese-stem content (no A_en counterpart)
     and B/C/D are filled. Move A_zh into stem_zh; shift options left.

  2. TF_FLASHCARD_IN_D: D_zh contains (T)/(F) flashcard lines that got
     appended. Strip them out.

  3. INLINE_LETTER_MARKER_BLEED: option_en has an inline 'B.' / 'C.' / 'D.'
     marker mid-text. Split on it, redistribute.

User-cleaned region (Q1-Q224) is preserved verbatim.
"""
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

ROOT = Path('/Users/gavincheung/NYU/Driver')
XLSX = ROOT / 'data' / 'questions_for_cleaning_v2.xlsx'
USER_DONE_THROUGH_ROW = 225

CJK_RE = re.compile(r'[一-鿿]')


def looks_like_stem_continuation(text):
    """True if text is mostly CJK and lacks substantial English."""
    if not text:
        return False
    cjk = len(CJK_RE.findall(text))
    ascii_letters = len(re.findall(r'[A-Za-z]', text))
    return cjk >= 5 and ascii_letters < 5


def strip_tf_flashcards(text):
    """Remove TF flashcard statements that got merged into option text.

    Two patterns:
      1. With marker: 'This is a XXX sign. ... (T)' — anywhere in text
      2. Bare:       'This is a XXX [road ]sign.' followed by Chinese — at end of text
    """
    if not text:
        return text, False
    original = text
    # Pattern 1: with explicit (T)/(F) marker
    pattern_marker = re.compile(
        r"(?:\s*This is an?\s[^\n]*?\(\s*[TF]\s*\)\s*)+",
        flags=re.IGNORECASE
    )
    text = pattern_marker.sub(' ', text)
    # Pattern 2: bare 'This is a/an ___ sign[.]' followed by optional CJK
    pattern_bare = re.compile(
        r"\s+This is an?\s.*?(?:sign|路标|标志)[\.。]?\s*(?:[一-鿿\"'’\s,，。\.\-/]*)\s*$",
        flags=re.IGNORECASE
    )
    text = pattern_bare.sub('', text)
    # Trailing solo markers
    text = re.sub(r'\s*\(\s*[TF]\s*\)\s*$', '', text)
    text = text.strip()
    return text, text != original


def split_inline_marker(en_text):
    """If en_text has inline ' B.'/'C.'/'D.' marker mid-text, split on it.
    Returns (head, rest_letter, rest_text) or (en_text, None, None)."""
    if not en_text:
        return en_text, None, None
    m = re.search(r'\s+([BCD])\.\s', en_text)
    if m:
        head = en_text[:m.start()].strip()
        letter = m.group(1)
        rest = en_text[m.end():].strip()
        return head, letter, rest
    return en_text, None, None


def main():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb['questions']
    headers = [c.value for c in ws[1]]
    H = {h: i for i, h in enumerate(headers)}

    fix_color = PatternFill('solid', fgColor='D1FAE5')  # green for systemic-fix
    review_color = PatternFill('solid', fgColor='FEF3C7')  # amber for needs review

    fixed_stem_shift = []
    fixed_tf_strip = []
    fixed_inline = []

    for row_idx in range(USER_DONE_THROUGH_ROW + 1, ws.max_row + 1):
        qid = ws[row_idx][H['id']].value
        if not qid:
            continue
        qid = int(qid)

        # Read current cell values (treating None as '')
        def get(col):
            v = ws[row_idx][H[col]].value
            return '' if v is None else str(v)

        def put(col, val):
            ws[row_idx][H[col]].value = val if val else None

        a_en = get('A_en'); a_zh = get('A_zh')
        b_en = get('B_en'); b_zh = get('B_zh')
        c_en = get('C_en'); c_zh = get('C_zh')
        d_en = get('D_en'); d_zh = get('D_zh')

        # ---- Fix 1: STEM_AS_OPTION_A ----
        # A_en empty, A_zh looks like stem continuation, B-filled
        if (not a_en.strip()
                and looks_like_stem_continuation(a_zh)
                and (b_en.strip() or b_zh.strip())):
            stem_zh = get('stem_zh')
            # Merge A_zh into stem_zh
            new_stem_zh = (stem_zh + ' ' + a_zh).strip() if stem_zh else a_zh.strip()
            put('stem_zh', new_stem_zh)
            # Shift B->A, C->B, D->C, leave D empty
            put('A_en', b_en); put('A_zh', b_zh)
            put('B_en', c_en); put('B_zh', c_zh)
            put('C_en', d_en); put('C_zh', d_zh)
            put('D_en', ''); put('D_zh', '')
            fixed_stem_shift.append(qid)
            # Mark as needs_review since D is now empty (likely needs source check)
            existing_nr = (ws[row_idx][H['needs_review']].value or '').strip()
            if 'orphan_answer' not in existing_nr:
                ws[row_idx][H['needs_review']].value = (existing_nr + ';' if existing_nr else '') + 'orphan_after_shift'
            for cell in ws[row_idx]:
                cell.fill = review_color
            # Re-read shifted values for subsequent fixes
            a_en = get('A_en'); a_zh = get('A_zh')
            b_en = get('B_en'); b_zh = get('B_zh')
            c_en = get('C_en'); c_zh = get('C_zh')
            d_en = get('D_en'); d_zh = get('D_zh')

        # ---- Fix 2: TF_FLASHCARD_IN_D and others ----
        any_tf_strip = False
        for col, val in [('A_zh', a_zh), ('B_zh', b_zh), ('C_zh', c_zh), ('D_zh', d_zh),
                         ('A_en', a_en), ('B_en', b_en), ('C_en', c_en), ('D_en', d_en)]:
            new_val, changed = strip_tf_flashcards(val)
            if changed:
                put(col, new_val)
                any_tf_strip = True
        if any_tf_strip:
            fixed_tf_strip.append(qid)
            if qid not in fixed_stem_shift:
                # Tint green for sanity-check (don't override review color)
                pass

        # ---- Fix 3: INLINE_LETTER_MARKER_BLEED ----
        # Re-read after fixes 1/2
        a_en = get('A_en'); b_en = get('B_en'); c_en = get('C_en'); d_en = get('D_en')
        b_zh = get('B_zh'); c_zh = get('C_zh'); d_zh = get('D_zh')
        # Check A_en for 'B.' bleed when B is empty
        head, rest_letter, rest_text = split_inline_marker(a_en)
        if rest_letter and not get('B_en').strip():
            put('A_en', head)
            put(f'{rest_letter}_en', rest_text)
            fixed_inline.append(qid)

    # ---- Fix 4: ALT-LINE BILINGUAL PAIRING ----
    # Each option's English and Chinese ended up in adjacent slots
    # Pattern: A_en filled+A_zh empty, B_en empty+B_zh filled => one option split
    fixed_alt = []
    for row_idx in range(USER_DONE_THROUGH_ROW + 1, ws.max_row + 1):
        qid = ws[row_idx][H['id']].value
        if not qid:
            continue
        qid = int(qid)

        def g(col):
            v = ws[row_idx][H[col]].value
            return '' if v is None else str(v).strip()

        def s(col, val):
            ws[row_idx][H[col]].value = val if val else None

        # Build the 4 (en, zh) pairs
        pairs = [(g(f'{L}_en'), g(f'{L}_zh')) for L in 'ABCD']

        # Detect: alternating en-only then zh-only pattern
        # Each pair is "en-only" if en filled but zh empty, "zh-only" if zh filled but en empty
        def kind(p):
            en, zh = p
            if en and not zh: return 'en'
            if zh and not en: return 'zh'
            if en and zh: return 'both'
            return 'empty'

        kinds = [kind(p) for p in pairs]
        # Look for [..., 'en', 'zh', ...] adjacent: merge them
        i = 0
        new_pairs = []
        merged_any = False
        while i < len(pairs):
            if i + 1 < len(pairs) and kinds[i] == 'en' and kinds[i + 1] == 'zh':
                new_pairs.append((pairs[i][0], pairs[i + 1][1]))
                merged_any = True
                i += 2
            else:
                new_pairs.append(pairs[i])
                i += 1
        # Pad to 4
        while len(new_pairs) < 4:
            new_pairs.append(('', ''))

        if merged_any:
            for k, L in enumerate('ABCD'):
                s(f'{L}_en', new_pairs[k][0])
                s(f'{L}_zh', new_pairs[k][1])
            fixed_alt.append(qid)
            # Mark needs_review for the orphan slots that came from merging
            empty_count = sum(1 for p in new_pairs if not p[0] and not p[1])
            if empty_count > 0:
                existing_nr = (ws[row_idx][H['needs_review']].value or '').strip()
                if 'orphan_after_pair_merge' not in existing_nr:
                    ws[row_idx][H['needs_review']].value = (existing_nr + ';' if existing_nr else '') + 'orphan_after_pair_merge'
            for cell in ws[row_idx]:
                cell.fill = review_color

    wb.save(XLSX)

    # ---- Fix 5: TRAILING OPTION IN X_ZH ----
    # An X_zh like "<CJK chunk> <long English run> <CJK chunk>" packs TWO
    # options' content. Split: first CJK = X's zh; English + remaining CJK = next slot.
    fixed_split = []
    # Find a long English run (next option's en) flanked by spaces and CJK.
    # Group 2 starts with [A-Z] then a letter (allowing 'A right' style options).
    # Group 1: capital letter, then any letter OR space (allows 'A right curve'),
    # then 8+ more chars (letters/digits/spaces/common punct).
    TRAILING_OPT_BOUNDARY = re.compile(
        r'\s+([A-Z][A-Za-z\s][A-Za-z\d\s/\.\',\-]{8,}?)\s+(?=[一-鿿])'
    )
    for row_idx in range(USER_DONE_THROUGH_ROW + 1, ws.max_row + 1):
        qid = ws[row_idx][H['id']].value
        if not qid:
            continue
        qid = int(qid)

        def g(c):
            v = ws[row_idx][H[c]].value
            return '' if v is None else str(v).strip()

        def s(c, v):
            ws[row_idx][H[c]].value = v if v else None

        # Loop until no more splits possible (Q251 has 2 trailing options)
        row_touched = False
        for _attempt in range(3):  # max 3 passes per row
            did_split = False
            for src_letter in ['C', 'B']:  # split FROM B and C only, into C or D
                src_zh_col = f'{src_letter}_zh'
                src_zh = g(src_zh_col)
                if not src_zh:
                    continue
                m = TRAILING_OPT_BOUNDARY.search(src_zh)
                if not m:
                    continue
                first_cjk = src_zh[:m.start()].strip()
                if not first_cjk or not re.match(r'[一-鿿]', first_cjk):
                    continue
                mid_en = m.group(1).strip()
                rest = src_zh[m.end():].strip()
                if not rest:
                    continue
                # Find empty slot to write extracted option
                dest_letter = None
                for L in 'BCD':
                    if L > src_letter:
                        if not g(f'{L}_en') and not g(f'{L}_zh'):
                            dest_letter = L
                            break
                if not dest_letter:
                    continue
                s(src_zh_col, first_cjk)
                s(f'{dest_letter}_en', mid_en)
                s(f'{dest_letter}_zh', rest)
                did_split = True
                row_touched = True
                break
            if not did_split:
                break
        if row_touched:
            fixed_split.append(qid)

        # Post-split: re-strip TF flashcards on the new D_zh
        for L in 'ABCD':
            col = f'{L}_zh'
            val = g(col)
            if val:
                new_val, changed = strip_tf_flashcards(val)
                if changed:
                    s(col, new_val)
                    fixed_tf_strip.append(qid)

    wb.save(XLSX)

    print(f'Saved {XLSX}')
    print(f'  stem-shift (A was stem cont):    {len(fixed_stem_shift)} questions')
    print(f'  TF flashcard stripped:           {len(fixed_tf_strip)} questions')
    print(f'  inline letter-marker split:      {len(fixed_inline)} questions')
    print(f'  alt-line bilingual pair merged:  {len(fixed_alt)} questions')
    print(f'  trailing-option split out:       {len(fixed_split)} questions')
    print(f'  Total unique IDs touched:        {len(set(fixed_stem_shift) | set(fixed_tf_strip) | set(fixed_inline) | set(fixed_alt) | set(fixed_split))}')
    print()
    if fixed_stem_shift[:10]:
        print(f'  Sample stem-shift IDs:  {fixed_stem_shift[:10]}')
    if fixed_alt[:10]:
        print(f'  Sample alt-merge IDs:   {fixed_alt[:10]}')


if __name__ == '__main__':
    main()
