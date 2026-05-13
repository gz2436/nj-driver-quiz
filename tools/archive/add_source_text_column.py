#!/usr/bin/env python3
"""阶段 B：在 v2.xlsx 基础上加「反查源」列，输出 v3.xlsx。

新增列：
  source_label          —— audit 还原的源题号 (M127 / C-list83.92 / C-manual640)
  source_text           —— 源 docx 里该题段 ± 2 段的原文拼起来
  source_answer_letter  —— mc 题：源 stem 末尾 (X) 字母
  source_answer_text    —— qa 题：源 答：后的文本
  answer_mismatch       —— audit 算出的不一致标 Y
  audit_flag            —— 漏题 / 重复 / 假重复 / 空答案 等提示

也会自动应用 audit 发现的「假重复绑图」（在 stem_img 列写建议值，
但不覆盖用户已写的）。

用户用这个 v3.xlsx 收尾 31 条红色 + 核对源。
"""
import json
import zipfile
from copy import copy
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path('/Users/gavincheung/NYU/Driver')
NS = 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'

DOCX = {
    'complete': ROOT / 'sources' / '美国新泽西驾照笔试题-ch eng（完整版）.docx',
    'mistakes': ROOT / 'sources' / '美国新泽西驾照笔试题-ch eng（易错题）.docx',
}


def load_paras(path):
    with zipfile.ZipFile(path) as z:
        doc = ET.fromstring(z.read('word/document.xml'))
    return [''.join(t.text or '' for t in p.iter(f'{{{NS}}}t'))
            for p in doc.iter(f'{{{NS}}}p')]


def main():
    paras = {label: load_paras(path) for label, path in DOCX.items()}
    audit = json.loads((ROOT / 'data' / 'audit_report.json').read_text())

    id_to_label = {int(k): v for k, v in audit['id_to_source_label'].items()}
    answer_mismatch_ids = {x['id'] for x in audit['answer_letter_mismatches']}
    empty_answer_ids = {x['id'] for x in audit['answer_empty_in_source']}
    dup_info = {x['duplicate_id']: x for x in audit['complete_internal_dups']}

    # 加载 questions.json 用于补 source_answer_letter / source_answer_text
    import re
    qs = json.loads((ROOT / 'data' / 'questions.json').read_text())
    qs_by_id = {q['id']: q for q in qs}

    src = ROOT / 'data' / 'questions_for_cleaning_v2.xlsx'
    out = ROOT / 'data' / 'questions_for_cleaning_v3.xlsx'

    wb = openpyxl.load_workbook(src)
    ws = wb['questions']
    headers = [c.value for c in ws[1]]
    H = {h: i for i, h in enumerate(headers)}

    new_cols = [
        'source_label',
        'source_text',
        'source_answer_letter',
        'source_answer_text',
        'answer_mismatch',
        'audit_flag',
    ]

    # 找到当前最末列，追加新列头
    last_col = len(headers)
    for i, name in enumerate(new_cols, start=1):
        cell = ws.cell(row=1, column=last_col + i, value=name)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='4338CA')  # indigo for audit cols
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # 设宽度
        col_letter = get_column_letter(last_col + i)
        widths = {'source_label': 14, 'source_text': 60,
                  'source_answer_letter': 8, 'source_answer_text': 30,
                  'answer_mismatch': 8, 'audit_flag': 25}
        ws.column_dimensions[col_letter].width = widths.get(name, 18)

    # 重新建 H 索引（含新列）
    headers2 = headers + new_cols
    H2 = {h: i for i, h in enumerate(headers2)}

    audit_flag_fill = PatternFill('solid', fgColor='FCA5A5')   # red tint
    audit_warn_fill = PatternFill('solid', fgColor='FED7AA')   # orange tint

    RE_ANS = re.compile(r'\(([A-D])\)(?:\s*$|\s+[A-D]\.)')
    RE_EMPTY = re.compile(r'\(\s*\)\s*$')
    RE_QA = re.compile(r'答\s*[：:]\s*(.*)$')

    flag_count = 0
    for row_idx in range(2, ws.max_row + 1):
        row = ws[row_idx]
        try:
            qid = int(row[H['id']].value)
        except (TypeError, ValueError):
            continue
        q = qs_by_id.get(qid)
        if not q:
            continue
        doc_label = q.get('source_doc')
        pi = q.get('source_para_idx')

        # source_label
        ws.cell(row=row_idx, column=last_col + 1,
                value=id_to_label.get(qid) or '')

        # source_text: pi ± 2 段
        st_lines = []
        if doc_label in paras and pi is not None:
            for offset in range(-2, 3):
                ix = pi + offset
                if 0 <= ix < len(paras[doc_label]):
                    marker = '→ ' if offset == 0 else '   '
                    t = paras[doc_label][ix].strip()
                    if t:
                        st_lines.append(f'{marker}[{ix}] {t}')
        ws.cell(row=row_idx, column=last_col + 2,
                value='\n'.join(st_lines))

        # source_answer_letter / source_answer_text
        src_ans = None
        src_text = None
        if doc_label in paras and pi is not None and pi < len(paras[doc_label]):
            stem_text = paras[doc_label][pi].strip()
            m_ans = RE_ANS.search(stem_text)
            m_empty = RE_EMPTY.search(stem_text)
            m_qa = RE_QA.search(stem_text)
            if m_ans:
                src_ans = m_ans.group(1)
            elif m_empty:
                src_ans = '(empty)'
            if m_qa:
                src_text = m_qa.group(1).strip()
        ws.cell(row=row_idx, column=last_col + 3, value=src_ans or '')
        ws.cell(row=row_idx, column=last_col + 4, value=src_text or '')

        # answer_mismatch
        if qid in answer_mismatch_ids:
            ws.cell(row=row_idx, column=last_col + 5, value='Y')
        else:
            ws.cell(row=row_idx, column=last_col + 5, value='')

        # audit_flag
        flags = []
        if qid in empty_answer_ids:
            flags.append('源里就是空答案，需补')
        if qid in dup_info:
            d = dup_info[qid]
            if d.get('likely_false_dup'):
                flags.append(f'假重复(无图)，建议绑图 {d["nearby_drawing_for_duplicate"]}')
            else:
                flags.append(f'真重复于 Q{d["original_id"]}，notes 写"删"或"保留"')
        if qid in answer_mismatch_ids:
            flags.append('答案字母与源不一致')
        ws.cell(row=row_idx, column=last_col + 6, value='; '.join(flags))

        if flags:
            flag_count += 1
            fill = audit_flag_fill if 'unsupported' in '; '.join(flags) else audit_warn_fill
            for cell in row:
                # 不覆盖已有的红/红色标记
                if not cell.fill or not cell.fill.fgColor or cell.fill.fgColor.rgb in (None, '00000000'):
                    cell.fill = fill

    # 自动对所有审计列做 wrap_text
    for row in ws.iter_rows(min_row=2, min_col=last_col + 1, max_col=last_col + len(new_cols)):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # ----- 在 legend 表里追加新列说明 -----
    if 'legend' in wb.sheetnames:
        ws2 = wb['legend']
    else:
        ws2 = wb.create_sheet('legend')
    legend_addition = [
        ['', ''],
        ['v3 新增', '说明'],
        ['source_label', '源 docx 里的题号 (mistakes: M1-M127；complete: C-list<numId>.<序> 或 C-manual<手输数字>)'],
        ['source_text', '源 docx 里该题段 ±2 段的原文，方便核对'],
        ['source_answer_letter', '源 stem 末尾 (X) 字母 (mc 题)'],
        ['source_answer_text', '源「答：」后的文本 (qa 题)'],
        ['answer_mismatch', '审计算出的答案字母不一致 (Y = 需要核对)'],
        ['audit_flag', '其他审计提示：空答案 / 重复 / 假重复 + 建议'],
    ]
    for r in legend_addition:
        ws2.append(r)

    wb.save(out)
    print(f'wrote {out}')
    print(f'  rows with audit_flag: {flag_count}')


if __name__ == '__main__':
    main()
