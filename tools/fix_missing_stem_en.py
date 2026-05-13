#!/usr/bin/env python3
"""Find rows where stem_en is empty but stem_zh is present, look back in the
source docx for the missing English stem, and fill it in.

Operates on data/questions_for_cleaning_v2.xlsx in-place.
"""
import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl

NS_W = 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'
ROOT = Path('/Users/gavincheung/NYU/Driver')
XLSX = ROOT / 'data' / 'questions_for_cleaning_v2.xlsx'

DOCX = {
    'complete': ROOT / 'sources' / '美国新泽西驾照笔试题-ch eng（完整版）.docx',
    'mistakes': ROOT / 'sources' / '美国新泽西驾照笔试题-ch eng（易错题）.docx',
}

RE_ANSWER = re.compile(r'\(([A-D])\)\s*$|\(\s*\)\s*$')
RE_QA = re.compile(r'答\s*[：:]')
RE_OPTION_LETTER = re.compile(r'(?:^|(?<=[^A-Z]))[ABCD][.．。]\s*')


def read_paras(path):
    with zipfile.ZipFile(path) as z:
        return [''.join(t.text or '' for t in p.iter(f'{{{NS_W}}}t'))
                for p in ET.fromstring(z.read('word/document.xml')).iter(f'{{{NS_W}}}p')]


def lookback_for_english_stem(paras, pi, max_look=5):
    """Walk back from paras[pi] to find the English half of the stem.

    Allow skipping ONE CJK-only paragraph (the Chinese stem, which is already
    captured as stem_zh). Stop at prior stems, Q&A answers, or option markers.
    """
    parts = []
    zh_skipped = False
    for k in range(pi - 1, max(-1, pi - 1 - max_look), -1):
        text = paras[k].strip()
        if not text:
            continue
        if RE_ANSWER.search(text):
            break
        if RE_QA.search(text):
            break
        if RE_OPTION_LETTER.search(text):
            break
        cjk = len(re.findall(r'[一-鿿]', text))
        ascii_letters = len(re.findall(r'[A-Za-z]', text))
        if cjk > 5 and ascii_letters < 5:
            # Pure CJK paragraph — skip once (Chinese stem half), then stop
            if not zh_skipped and not parts:
                zh_skipped = True
                continue
            break
        parts.insert(0, text)
        if sum(len(p) for p in parts) > 30 and any(re.search(r'[A-Za-z]', p) for p in parts):
            break
    if not parts:
        return None
    joined = ' '.join(parts).strip()
    if len(re.findall(r'[A-Za-z]', joined)) < 8:
        return None
    return joined


def main():
    docs = {k: read_paras(v) for k, v in DOCX.items()}

    wb = openpyxl.load_workbook(XLSX)
    ws = wb['questions']
    headers = [c.value for c in ws[1]]
    H = {h: i for i, h in enumerate(headers)}

    fixed = []
    skipped_no_lookback = []

    for row_idx in range(2, ws.max_row + 1):
        stem_en = (ws[row_idx][H['stem_en']].value or '').strip()
        stem_zh = (ws[row_idx][H['stem_zh']].value or '').strip()
        if stem_en:
            continue
        if not stem_zh:
            continue
        # Skip Q&A rows — they don't have English stems in source typically
        qtype = (ws[row_idx][H['type']].value or '').strip()
        if qtype == 'qa':
            continue
        # Skip statement rows
        if qtype == 'statement':
            continue
        # Skip user-cleaned region (rows 2..225) — leave their work alone
        if row_idx <= 225:
            continue

        qid = ws[row_idx][H['id']].value
        doc = ws[row_idx][H['source_doc']].value
        pi_val = ws[row_idx][H['source_para_idx']].value
        if not doc or pi_val is None or doc not in docs:
            continue
        try:
            pi = int(pi_val)
        except (ValueError, TypeError):
            continue

        candidate = lookback_for_english_stem(docs[doc], pi)
        if candidate:
            # Strip trailing ":" that's typically the English colon
            candidate = candidate.rstrip(' \t')
            ws[row_idx][H['stem_en']].value = candidate
            fixed.append((qid, candidate[:80]))
        else:
            skipped_no_lookback.append(qid)

    wb.save(XLSX)
    print(f'Fixed stem_en for {len(fixed)} rows')
    for qid, preview in fixed[:25]:
        print(f'  Q{qid}: {preview}')
    if len(fixed) > 25:
        print(f'  ... +{len(fixed) - 25} more')
    if skipped_no_lookback:
        print(f'\nSkipped (no English found within 4 paragraphs back): {len(skipped_no_lookback)}')
        print(f'  IDs: {skipped_no_lookback[:30]}')


if __name__ == '__main__':
    main()
