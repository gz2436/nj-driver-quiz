#!/usr/bin/env python3
"""阶段 D：把 data/questions_for_cleaning_v3.xlsx 合并回 data/questions.json。

行为：
  1. 备份原 questions.json -> questions.<ts>.backup.json
  2. 按 id 对齐 xlsx 行与 questions.json 题
  3. 字段合并规则：
     - stem_en / stem_zh / A_en / A_zh / ... 全部从 xlsx 覆盖原值
     - 同时生成 stem (= stem_en + " " + stem_zh) 和 options[]（合并双语）做向后兼容
     - answer / answer_text / topics / is_common_mistake / explanation_key 从 xlsx 覆盖
     - is_duplicate_of / needs_review 从 xlsx 覆盖
     - verified: xlsx 里若有 Y 则 True
     - stem_img: 若 xlsx 空但 audit_report 建议绑图，则用建议值
  4. notes 列含 "删" / "delete" / "drop" → 标 dropped=true (不真删)
  5. 加 lang_primary / lang_secondary / needs_translation 字段
  6. 补 3 漏题（题+答+选项同一行的格式）id 续编
  7. 写回 data/questions.json
"""
import datetime as dt
import json
import re
import shutil
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl

ROOT = Path('/Users/gavincheung/NYU/Driver')
NS = 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'

DOCX = {
    'complete': ROOT / 'sources' / '美国新泽西驾照笔试题-ch eng（完整版）.docx',
    'mistakes': ROOT / 'sources' / '美国新泽西驾照笔试题-ch eng（易错题）.docx',
}

# 漏题（来自 audit 报告），格式: "stem (X) A. xB. xC. xD. x"
MISSING_QUESTIONS = [
    {'source_doc': 'complete', 'source_para_idx': 36,  'is_common_mistake': True},  # 也在 mistakes
    {'source_doc': 'complete', 'source_para_idx': 527},
    {'source_doc': 'mistakes', 'source_para_idx': 7,
     'is_common_mistake': True, 'is_duplicate_of_complete_pi': 36},
]

# 单行题正则: "<stem> (X) A. <a>B. <b>C. <c>D. <d>"
SINGLE_LINE_Q = re.compile(
    r'^(?P<stem>.+?)\((?P<ans>[A-D])\)\s*A[.．]\s*(?P<a>.*?)B[.．]\s*'
    r'(?P<b>.*?)C[.．]\s*(?P<c>.*?)D[.．]\s*(?P<d>.+)$'
)

CJK_RE = re.compile(r'[一-鿿]')


def split_bilingual(text):
    if not text:
        return '', ''
    m = CJK_RE.search(text)
    if not m:
        return text.strip(), ''
    return text[:m.start()].strip(), text[m.start():].strip()


def join_bilingual(en, zh):
    en = (en or '').strip()
    zh = (zh or '').strip()
    if en and zh:
        return f'{en} {zh}'
    return en or zh


def load_paras(path):
    with zipfile.ZipFile(path) as z:
        doc = ET.fromstring(z.read('word/document.xml'))
    return [''.join(t.text or '' for t in p.iter(f'{{{NS}}}t'))
            for p in doc.iter(f'{{{NS}}}p')]


def parse_single_line_question(text):
    """Parse 'stem (X) A. xB. xC. xD. x' format."""
    m = SINGLE_LINE_Q.search(text)
    if not m:
        return None
    d = m.groupdict()
    return {
        'stem': d['stem'].strip().rstrip(':：'),
        'answer': d['ans'],
        'options': [d['a'].strip(), d['b'].strip(), d['c'].strip(), d['d'].strip()],
    }


def determine_lang(stem_en, stem_zh):
    """Return (lang_primary, needs_translation)."""
    has_en = bool((stem_en or '').strip())
    has_zh = bool((stem_zh or '').strip())
    if has_en and has_zh:
        return 'en', None
    if has_en and not has_zh:
        return 'en', 'en_to_zh'
    if not has_en and has_zh:
        return 'zh', 'zh_to_en'
    return 'en', None  # both empty — leave as en


def xlsx_to_dict(ws, headers):
    H = {h: i for i, h in enumerate(headers)}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[H['id']] is None:
            continue
        rows.append({h: row[H[h]] for h in headers})
    return rows


def merge_row(xlsx_row, q, audit):
    """Update q in-place from xlsx_row."""
    g = lambda k: (xlsx_row.get(k) or '').strip() if isinstance(xlsx_row.get(k), str) else xlsx_row.get(k)

    stem_en = g('stem_en') or ''
    stem_zh = g('stem_zh') or ''
    q['stem_en'] = stem_en
    q['stem_zh'] = stem_zh
    q['stem'] = join_bilingual(stem_en, stem_zh)

    opts_en = [g(f'{L}_en') or '' for L in 'ABCD']
    opts_zh = [g(f'{L}_zh') or '' for L in 'ABCD']
    q['options_en'] = opts_en
    q['options_zh'] = opts_zh
    q['options'] = [join_bilingual(e, z) for e, z in zip(opts_en, opts_zh)]

    ans = g('answer')
    q['answer'] = (ans or '').strip().upper() or None

    at_en = g('answer_text_en') or ''
    at_zh = g('answer_text_zh') or ''
    q['answer_text_en'] = at_en
    q['answer_text_zh'] = at_zh
    q['answer_text'] = join_bilingual(at_en, at_zh) or None

    # 简单字段
    if g('stem_img'):
        q['stem_img'] = g('stem_img')
    elif xlsx_row['id'] in audit['suggest_img']:
        # 应用 audit 建议绑图
        q['stem_img'] = audit['suggest_img'][xlsx_row['id']]

    if g('topics'):
        q['topics'] = [t.strip() for t in str(g('topics')).split(',') if t.strip()]

    cm = g('is_common_mistake')
    q['is_common_mistake'] = cm in ('Y', 'y', True, 1)

    if g('is_duplicate_of'):
        try:
            q['is_duplicate_of'] = int(g('is_duplicate_of'))
        except (ValueError, TypeError):
            q['is_duplicate_of'] = None
    else:
        q['is_duplicate_of'] = None

    nr = g('needs_review')
    q['needs_review'] = nr if nr else None

    q['verified'] = g('verified') in ('Y', 'y', True, 1)

    if g('explanation_key'):
        q['explanation_key'] = g('explanation_key')

    # 双语策略字段
    primary, needs_t = determine_lang(stem_en, stem_zh)
    q['lang_primary'] = primary
    q['needs_translation'] = needs_t

    # notes 列指令
    notes = (g('notes') or '').strip()
    drop_markers = ('删', 'delete', 'drop', '删除')
    if any(m in notes for m in drop_markers):
        q['dropped'] = True
    q['notes'] = notes if notes else None


def main():
    src_xlsx = ROOT / 'data' / 'questions_for_cleaning_v3.xlsx'
    src_json = ROOT / 'data' / 'questions.json'
    audit_path = ROOT / 'data' / 'audit_report.json'

    if not src_xlsx.exists():
        print(f'ERROR: {src_xlsx} not found. Run tools/add_source_text_column.py first.')
        return

    # 备份原文件
    ts = dt.datetime.now().strftime('%Y%m%d_%H%M%S')
    backup = src_json.with_suffix(f'.{ts}.backup.json')
    shutil.copy2(src_json, backup)
    print(f'backed up to {backup.name}')

    qs = json.loads(src_json.read_text())
    qs_by_id = {q['id']: q for q in qs}

    # 读 audit 报告 + media 映射
    audit_data = json.loads(audit_path.read_text()) if audit_path.exists() else {}
    media_map_path = ROOT / 'data' / 'media_to_local.json'
    media_map = json.loads(media_map_path.read_text()) if media_map_path.exists() else {}

    def media_to_local(media_path):
        """'media/image9.jpg' -> 'img_complete_001.jpg' (bare filename for front-end)"""
        short = media_path.removeprefix('media/') if media_path.startswith('media/') else media_path
        for key in [f'complete/{short}', f'mistakes/{short}']:
            if key in media_map:
                return media_map[key]
        return None

    suggest_img = {}
    for d in audit_data.get('complete_internal_dups', []):
        if d.get('likely_false_dup') and d.get('nearby_drawing_for_duplicate'):
            local = media_to_local(d['nearby_drawing_for_duplicate'])
            if local:
                suggest_img[d['duplicate_id']] = local
    audit_compact = {'suggest_img': suggest_img}
    print(f'  audit-suggested image bindings: {suggest_img}')

    # 合并
    wb = openpyxl.load_workbook(src_xlsx, read_only=True, data_only=True)
    ws = wb['questions']
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = xlsx_to_dict(ws, headers)

    merged = 0
    dropped = 0
    suggested_imgs = 0
    covered = set()
    for row in rows:
        qid = int(row['id'])
        covered.add(qid)
        q = qs_by_id.get(qid)
        if not q:
            continue
        before_drop = q.get('dropped')
        before_img = q.get('stem_img')
        merge_row(row, q, audit_compact)
        merged += 1
        if q.get('dropped') and not before_drop:
            dropped += 1
        if q.get('stem_img') and not before_img:
            suggested_imgs += 1

    # 对 xlsx 没覆盖到的 questions，从 stem 字段自动拆双语 + 算 lang_primary
    uncovered = []
    for q in qs:
        if q['id'] in covered:
            continue
        stem_en, stem_zh = split_bilingual(q.get('stem') or '')
        q['stem_en'] = stem_en
        q['stem_zh'] = stem_zh
        opts_split = [split_bilingual(o) for o in (q.get('options') or [])]
        q['options_en'] = [e for e, _ in opts_split]
        q['options_zh'] = [z for _, z in opts_split]
        at_en, at_zh = split_bilingual(q.get('answer_text') or '')
        q['answer_text_en'] = at_en
        q['answer_text_zh'] = at_zh
        primary, needs_t = determine_lang(stem_en, stem_zh)
        q['lang_primary'] = primary
        q['needs_translation'] = needs_t
        uncovered.append(q['id'])
    if uncovered:
        print(f'  uncovered by xlsx (auto-split bilingual): {len(uncovered)} IDs={uncovered[:10]}{"..." if len(uncovered)>10 else ""}')

    # ---- 补漏题（按 source_doc + pi 去重，幂等可重跑） ----
    paras = {label: load_paras(path) for label, path in DOCX.items()}
    existing_pi = {(q.get('source_doc'), q.get('source_para_idx')) for q in qs}
    next_id = max(q['id'] for q in qs) + 1
    added = []
    for spec in MISSING_QUESTIONS:
        key = (spec['source_doc'], spec['source_para_idx'])
        if key in existing_pi:
            print(f'  missing question at {key} already present — skip')
            continue
        text = paras[spec['source_doc']][spec['source_para_idx']].strip()
        parsed = parse_single_line_question(text)
        if not parsed:
            print(f'  WARN: failed to parse missing question at {spec}')
            continue
        stem_en, stem_zh = split_bilingual(parsed['stem'])
        opts_split = [split_bilingual(o) for o in parsed['options']]
        opts_en = [e for e, _ in opts_split]
        opts_zh = [z for _, z in opts_split]
        primary, needs_t = determine_lang(stem_en, stem_zh)
        new_q = {
            'id': next_id,
            'type': 'mc',
            'stem': parsed['stem'],
            'stem_en': stem_en,
            'stem_zh': stem_zh,
            'stem_img': None,
            'options': parsed['options'],
            'options_en': opts_en,
            'options_zh': opts_zh,
            'answer': parsed['answer'],
            'answer_text': None,
            'answer_text_en': '',
            'answer_text_zh': '',
            'topics': ['unclassified'],
            'is_common_mistake': spec.get('is_common_mistake', False),
            'explanation_key': None,
            'verified': False,
            'needs_review': None,
            'source_doc': spec['source_doc'],
            'source_para_idx': spec['source_para_idx'],
            'is_duplicate_of': None,
            'lang_primary': primary,
            'needs_translation': needs_t,
            'notes': 'auto-added from audit (single-line stem+options+answer format)',
        }
        # 处理 mistakes 那条对应 complete 的引用（用 stem 等价做后置链接）
        if 'is_duplicate_of_complete_pi' in spec:
            target_pi = spec['is_duplicate_of_complete_pi']
            target_q = next((x for x in qs if x.get('source_doc') == 'complete'
                             and x.get('source_para_idx') == target_pi), None)
            if target_q is None:
                # 可能是新补的 3 道里的另一条
                target_q = next((a for a in added
                                 if a.get('source_doc') == 'complete'
                                 and a.get('source_para_idx') == target_pi), None)
            if target_q:
                new_q['is_duplicate_of'] = target_q['id']
        qs.append(new_q)
        added.append(new_q)
        next_id += 1

    src_json.write_text(json.dumps(qs, ensure_ascii=False, indent=2))
    print(f'  wrote {src_json}')
    print(f'  merged xlsx rows: {merged}')
    print(f'  dropped (marked): {dropped}')
    print(f'  auto-bound images: {suggested_imgs}')
    print(f'  appended missing questions: {len(added)}  IDs: {[a["id"] for a in added]}')
    print(f'  total now: {len(qs)}')

    # 验证
    from collections import Counter
    print()
    print('--- verification ---')
    print(f'  by type: {dict(Counter(q["type"] for q in qs))}')
    print(f'  with answer: {sum(1 for q in qs if q.get("answer"))}')
    print(f'  is_common_mistake: {sum(1 for q in qs if q.get("is_common_mistake"))}')
    print(f'  lang_primary breakdown: {dict(Counter(q.get("lang_primary") for q in qs))}')
    print(f'  needs_translation: {dict(Counter(q.get("needs_translation") for q in qs))}')
    print(f'  dropped: {sum(1 for q in qs if q.get("dropped"))}')
    print(f'  verified: {sum(1 for q in qs if q.get("verified"))}')


if __name__ == '__main__':
    main()
